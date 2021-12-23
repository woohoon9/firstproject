from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()
ws.title ="MySheet"
ws.sheet_properties.tabColor="ff66ff"

ws1=wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2=wb.create_sheet("NewSheet",2) # 2번째 index에 Sheet 생성
new_ws = wb["NewSheet"]
print(wb.sheetnames) # 모든 sheet 이름 확인

new_ws["A1"]="test"

target = wb.copy_worksheet(new_ws)
target.title = "Copied_Sheet"

wb.save("sample.xlsx")