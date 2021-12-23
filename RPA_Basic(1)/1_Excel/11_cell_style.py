from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

a1=ws["A1"]
b1=ws["B1"]
c1=ws["C1"]

ws.column_dimensions["A"].width = 5
ws.row_dimensions[1].height = 50

# Applying Style
a1.font = Font(color = "FF0000", italic=True, bold=True)
b1.font = Font(color = "CC33FF", name ="Arial", strike=True ) # Strike = 취소선
c1.font = Font(color = "0000FF", size=20, underline = "single")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

#  셀 배경 색깔 변경
for row in ws.rows:
    for cell in row:
#각 cell에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical ="center")

    # center, left, right, top, bottom

        if cell.column == 1:
            continue
# cell이 정수형 데이터이고 90점 보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") #  배경 초록색 설정
            cell.font = Font(color ="FF0000")


# 틀 고정

ws.freeze_panes = "B2"

wb.save("sample_style.xlsx")