import win32com.client as win32
import os
import pandas as pd
from openpyxl import Workbook

xlApp = win32.Dispatch('Excel.Application')
xlApp.Visible = True

#workbook creating and sheet creating
# wb = xlApp.Workbooks.Add()
# wb.SaveAs(os.path.join(os.getcwd(), 'text.xlsx'))
# ws_sheet1 = wb.Worksheets('Sheet1')
# ws_sheet1.name
# ws_sheet1.name = 'Dummy Data Test'


# xlsm1 = Workbook(r'C:\Wooswork\02. Automation Estimating\03_Lots Sample\Dynamo Python Scripts\Test\Change detection\Change Detection V01 xlsx.xlsx')
# xlsm1 = pd.ExcelFile (r'C:\Wooswork\02. Automation Estimating\03_Lots Sample\Dynamo Python Scripts\Test\Change detection\Change Detection V01 xlsx.xlsx')
# df1_0 = pd.read_excel(xlsm1, "00-CoverPage")
# df1_1 = pd.read_excel(xlsm1, "01.EST_Doors")
# df1_2 = pd.read_excel(xlsm1, "02.EST_Windows")
# df1_3 = pd.read_excel(xlsm1, "03.EST_STUD_Wall Types")
# df1_4 = pd.read_excel(xlsm1, "04.Foundation")

# print(df1_1.values)

wb = xlApp.Workbooks.Open(r'C:\Wooswork\02. Automation Estimating\03_Lots Sample\Dynamo Python Scripts\Test\Change detection\Change Detection V01 xlsx.xlsx')
df1_1= xlApp.Sheets('02.EST_Windows').select() # this just let the file and sheet pop up on the screen
#df1_1.Cells(5, "B").Value = "Cell B5" # It doesn't work
#df1_1.Cells(20, "C").Interior.ColorIndex = 3  # It doesn't work

# with pd.ExcelWriter(r"C:\Wooswork\02. Automation Estimating\03_Lots Sample\Dynamo Python Scripts\Test\Change detection\Change Detection V01 xlsx.xlsx", mode="a", engine="openpyxl") as writer: df1_1.to_excel(writer, sheet_name="01.EST_Doors")
#cell wrtings
# ws_sheet3.Cells(5, "B").Value = "Cell B5"
# ws_sheet3.Cells(5, "C").Value = "Cell C5"

# ws_sheet3.Range("D5").Value = "Cell D5"

# #cell color changing
# ws_sheet3.Cells(20, "C").Interior.ColorIndex = 3

# ws_sheet3.Range("B15").Interior.ColorIndex = 12

#Cells (row index, column index) #row index, column indes

# cell wrtings
# ws_sheet1.Cells(5, "B").Value = "Cell B5"
# ws_sheet1.Cells(5, "C").Value = "Cell C5"

# ws_sheet1.Range("D5").Value = "Cell D5"

# #cell color changing
# ws_sheet1.Cells(20, "C").Interior.ColorIndex = 3

# ws_sheet1.Range("B15").Interior.ColorIndex = 12


# """ import os
# from openpyxl import load_workbook

# def append_df_to_excel(filename, df2_1, sheet_name='01.EST_Doors', startrow=None,
# #                        truncate_sheet=False, 
# #                        **to_excel_kwargs):
# #     """

    # """Append a DataFrame [df] to existing Excel file [filename]
    # into [sheet_name] Sheet.
    # If [filename] doesn't exist, then this function will create it.

    # @param filename: File path or existing ExcelWriter
    #                  (Example: '/path/to/file.xlsx')
    # @param df: DataFrame to save to workbook
    # @param sheet_name: Name of sheet which will contain DataFrame.
    #                    (default: 'Sheet1')
    # @param startrow: upper left cell row to dump data frame.
    #                  Per default (startrow=None) calculate the last row
    #                  in the existing DF and write to the next row...
    # @param truncate_sheet: truncate (remove and recreate) [sheet_name]
    #                        before writing DataFrame to Excel file
    # @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
    #                         [can be a dictionary]
    # @return: None

    # Usage examples:

    # >>> append_df_to_excel('d:/temp/test.xlsx', df)

    # >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    # >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
    #                        index=False)

    # >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
    #                        index=False, startrow=25)

    # (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    # """
    # """# Excel file doesn't exist - saving and exiting
    # if not os.path.isfile(filename):
    #     df.to_excel(
    #         filename,
    #         sheet_name=sheet_name, 
    #         startrow=startrow if startrow is not None else 0, 
    #         **to_excel_kwargs)
    #     return
    
    # # ignore [engine] parameter if it was passed
    # if 'engine' in to_excel_kwargs:
    #     to_excel_kwargs.pop('engine')

    # writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # # try to open an existing workbook
    # writer.book = load_workbook(filename)
    
    # # get the last row in the existing Excel sheet
    # # if it was not specified explicitly
    # if startrow is None and sheet_name in writer.book.sheetnames:
    #     startrow = writer.book[sheet_name].max_row

    # # truncate sheet
    # if truncate_sheet and sheet_name in writer.book.sheetnames:
    #     # index of [sheet_name] sheet
    #     idx = writer.book.sheetnames.index(sheet_name)
    #     # remove [sheet_name]
    #     writer.book.remove(writer.book.worksheets[idx])
    #     # create an empty sheet [sheet_name] using old index
    #     writer.book.create_sheet(sheet_name, idx)
    
    # # copy existing sheets
    # # writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    # if startrow is None:
    #     startrow = 0

    # # write out the new sheet
    # df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # # save the workbook
    # writer.save() """
